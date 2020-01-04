import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patheffects as PathEffects
# %matplotlib inline

from sklearn.manifold import TSNE

import seaborn as sns
from openpyxl import load_workbook

sns.set_style('darkgrid')
sns.set_palette('muted')
sns.set_context("notebook", font_scale=1.5,
                rc={"lines.linewidth": 2.5})
RS = 1

# get indices of the pictures from the cell value that looks like "bs(2)bs(3)" and converts it to - 2,3

# maps between person names (as they are in the Excel file) to the folder & pic names (initials)
map_name_to_folder = {"Blanca":"BS", "Franka":"FP", "Giovanna":"GM","Johanna":"JS","Noomi":"NR","Carlos":"CL","Francesco":"FM","Guillame":"GC","Lambert":"LW","Stefano":"SA"}


def transform_indices(cell):
    # print(cell)
    splitted_cell = cell.split("(")

    # print(splitted_cell)
    first_picture = splitted_cell[1].split(")")[0]
    second_picture = splitted_cell[2].split(")")[0]

    return first_picture,second_picture


names = ["Blanca", "Franka", "Giovanna", "Johanna","Noomi","Carlos","Francesco","Guillame","Lambert","Stefano"]


def resize_images_func(size):
    from PIL import Image
    import os,sys

    for name in names:
        path = f"faces_for_clustering/{map_name_to_folder[name]}"
        dirs = os.listdir( path )

        for item in dirs:
            if os.path.isfile(f"{path}/{item}"):
                im = Image.open(f"{path}/{item}")
                f, e = os.path.splitext(f"{path}/{item}")
                imResize = im.resize((size,size), Image.ANTIALIAS)
                imResize.save(f"{f}.jpg", 'JPEG', quality=90)
        print(f"Done with name {name}")
    print("DONE")

# resize_images_func(100)


# reads the excel and parses it to a collection of distance matrices for each person
# returns two maps:
#    person's name to distance matrix
#    person's name to pics names to index (array)

def create_dist_mat():
    wb = load_workbook(filename = 'Similarity_Ratings_human_fixed.xlsx')
    # names = ["Blanca"]
    total_matrices = {}
    total_names_to_indices= {}
    for name in names:
        names_to_index = []
        dist_matrix = [[0]*15 for i in range(15)]  # init an empty distance matrix of size 15*15
        first = True
        for row in wb[name].rows: # for each row (skip the first row)
            if first:
                first = False
                continue
            pic1, pic2 = transform_indices(row[0].value)
            # init a name to index mapping (an array of size 15) for each person
            if pic1 not in names_to_index:
                names_to_index.append(pic1)
            if pic2 not in names_to_index:
                names_to_index.append(pic2)
            ind1 = names_to_index.index(pic1)
            ind2 = names_to_index.index(pic2)
            dist_matrix[ind1][ind2] = row[1].value # set the distance between pic1 and pic2 in the dist matrix
            dist_matrix[ind2][ind1] = row[1].value # same
        # print(dist_matrix)
        total_matrices[name] = dist_matrix # set the mapping from person name to dist matrix
        total_names_to_indices[name] = names_to_index # set the mapping from person name to his pics to index array
    return total_matrices, total_names_to_indices


# distance_matrices, name_to_picture_names = create_dist_mat()

names_algo = ["Blanca", "Carlos", "Francesco", "Franka", "Giovanna", "Guillame", "Johanna", "Lambert", "Noomi", "Stefano"]


def create_pic_names_array(row):
    for i in range(150)

def create_dist_mat_algo():
    wb = load_workbook(filename = 'final_distances_matrix .xlsx')
    # names = ["Blanca"]
    total_matrices = {}
    total_names_to_indices= {}
    row_counter = 0
    total_matrices= {}
    for name in names_algo:
        total_matrices[name] =[[0]*15 for i in range(15)]
    for row in wb["final_distances_matrix "].rows:
        if row_counter==1:
            map_names_to_pic_names = create_pic_names_array(row)

        if row_counter>=2:
            current_name = (row_counter -2)//15
            dist_matrix = total_matrices[current_name]
            dist_matrix[]


        row_counter+=1

    for name in names_algo:
        # indices [ 2-16] = 15*index(name)+2-15*index(Name)+16
        # 15*indexname ( 2 to 16)

        names_to_index = []
          # init an empty distance matrix of size 15*15
        first = True
        for row in wb[name].rows: # for each row (skip the first row)

            # if first:
            #     first = False
            #     continue
            # pic1, pic2 = transform_indices(row[0].value)
            # # init a name to index mapping (an array of size 15) for each person
            # if pic1 not in names_to_index:
            #     names_to_index.append(pic1)
            # if pic2 not in names_to_index:
            #     names_to_index.append(pic2)
            # ind1 = names_to_index.index(pic1)
            # ind2 = names_to_index.index(pic2)
            # dist_matrix[ind1][ind2] = row[1].value # set the distance between pic1 and pic2 in the dist matrix
            # dist_matrix[ind2][ind1] = row[1].value # same
        # print(dist_matrix)
        total_matrices[name] = dist_matrix # set the mapping from person name to dist matrix
        total_names_to_indices[name] = names_to_index # set the mapping from person name to his pics to index array
    return total_matrices, total_names_to_indices


distance_matrices, name_to_picture_names = create_dist_mat()



# a plotting function to plot the result of TSNE dimensional reduction
# inputs:
#       x =  a two dimensional array, X[:,0] is x_axis, X[:,1] is y_axis
#       picture_names = the names of the pictures that correspond to each x,y point in the plot
#       name_person = a specific name of a person (Blanca, Carlos, etc)
def tsne_scatter(x, picture_names, name_person):
    # imports relevant for plotting an image in the graph
    from matplotlib.offsetbox import TextArea, DrawingArea, OffsetImage, AnnotationBbox
    import matplotlib.image as mpimg

    # choose a color palette with seaborn.
    # num_classes = len(np.unique(picture_names))
    # labels = picture_names

    # create a scatter plot.
    f = plt.figure(figsize=(8, 8))
    ax = plt.subplot(aspect='equal')
    sc = ax.scatter(x[:,0], x[:,1], lw=0, s=40)
    ax.axis('on')
    xmin, xmax, ymin, ymax = -500, 500, -500, 500  # TODO: check if these limits are enough
    ax.set(xlim=(xmin,xmax), ylim=(ymin,ymax))

    for i, pic_name in enumerate(picture_names):
        # ax.annotate(s=pic_name,xy=(x[:,0][i], x[:,1][i]))
        image_folder_path = 'faces_for_clustering'
        person_folder_pic_name = map_name_to_folder[name_person]
        image_path = f'{image_folder_path}/{person_folder_pic_name}/{person_folder_pic_name} ({picture_names[i]}).jpg'

        arr_img = mpimg.imread(image_path)  # open the image

        imagebox = OffsetImage(arr_img, zoom=0.3)  # create an image box with a certain zoom

        ab = AnnotationBbox(imagebox, xy=(x[:,0][i], x[:,1][i]), pad=0.1)  # create an annotation box, which is at XY on graph

        ax.add_artist(ab)

    plt.savefig(f"{name_person}.png")  # save the figure to the current directory

    plt.show()  # opens a window with the plot
    return (f, ax, sc)





# a main function to iterate over all persons and plot the tsne visualization

def main():
    for person_name in names:
        distance_np = np.array(distance_matrices[person_name ])
        print("------------------CALCULATING DISTANCE MATRIX----------------")
        print(distance_matrices[person_name])
        print("-------------------------------------------------------------")
        names_of_pics = np.array(name_to_picture_names[person_name ])
        print("------------------NAMES OF PICTURES----------------")
        print(names_of_pics)
        print("-------------------------------------------------------------")

        # perplexity is 2 because it showed the best clustering results
        tsne_object = TSNE(verbose=20, method= "exact", metric="precomputed", random_state=RS, perplexity=2)
        fashion_tsne = tsne_object.fit_transform(distance_matrices[person_name])
        print("-----------------------TSNE RESULT [0] -------------------------")
        print(fashion_tsne[:,0])
        print("-------------------------------------------------------------")
        print("-----------------------TSNE RESULT [1] -------------------------")
        print(fashion_tsne[:, 1])
        print("-------------------------------------------------------------")

        tsne_scatter(fashion_tsne, names_of_pics, person_name )

main()